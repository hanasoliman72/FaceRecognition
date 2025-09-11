import os
import cv2
import json
import numpy as np
import tkinter as tk
import face_recognition
from datetime import datetime
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

ENCODINGS_FILE = "registered_faces.json"
LOG_FILE = "login_log.xlsx"

# Load existing data
if os.path.exists(ENCODINGS_FILE):
    with open(ENCODINGS_FILE, "r") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            print("⚠️ Warning: JSON decode error. Starting with empty data.")
            data = []
else:
    data = []

#Save to Excel sheet
def log_login(name):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if name == "Hana" or name == "Nancy":
        is_team_member = "Yes"
    else: is_team_member = "No"

    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Login Time", "Team Member"])
    else:
        wb = load_workbook(LOG_FILE)
        ws = wb.active

    ws.append([name, now, is_team_member])
    wb.save(LOG_FILE)

def register_user(name):
    if not name:
        messagebox.showwarning("Input Error", "Please enter a name.")
        return

    # Check if name already exists
    if any(d.get("name", "").lower() == name.lower() for d in data):
        messagebox.showwarning("Duplicate Name", f"Name '{name}' already exists.")
        return

    # Start camera
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Camera Error", "Failed to access the camera.")
        return

    captured_encoding = None
    messagebox.showinfo("Face Capture", "Please look at the camera to capture your face.")

    while True:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Camera Error", "Failed to grab frame.")
            break

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb)
        face_encodings = face_recognition.face_encodings(rgb, face_locations)

        for top, right, bottom, left in face_locations:
            cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)

        cv2.imshow("Register - Press 's' to save, 'q' to quit", frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('s') and face_encodings:
            captured_encoding = face_encodings[0]

            # Check if this face already exists
            for user in data:
                existing_encoding = np.array(user["encoding"])
                distance = np.linalg.norm(existing_encoding - captured_encoding)
                if distance < 0.45:
                    messagebox.showwarning("Duplicate Face", "This face is already registered.")
                    cap.release()
                    cv2.destroyAllWindows()
                    return

            messagebox.showinfo("Face Captured", "Face captured successfully.")
            break
        elif key == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    if captured_encoding is not None:
        data.append({
            "name": name,
            "encoding": captured_encoding.tolist()
        })
        with open(ENCODINGS_FILE, "w") as f:
            json.dump(data, f, indent=2)
        messagebox.showinfo("Registration Successful", f"Successfully Registered {name}!")
    else:
        messagebox.showwarning("No Face Captured", "No face was captured during registration.")

def login_user(name):
    if not name:
        messagebox.showwarning("Input Error", "Please enter a name.")
        return

    if not os.path.exists(ENCODINGS_FILE):
        messagebox.showwarning("No Users Registered", "No registered users found.")
        return

    with open(ENCODINGS_FILE, "r") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            messagebox.showwarning("Error", "Error loading user data.")
            return

    known_names = [d.get("name", "") for d in data]
    known_encodings = [np.array(d["encoding"]) for d in data]

    if name.lower() not in [n.lower() for n in known_names]:
        messagebox.showwarning("Login Failed", f"No user found with the name '{name}'.")
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Camera Error", "Could not open webcam.")
        return

    messagebox.showinfo("Face Login", "Please look at the camera.")
    already_logged_in = False

    while True:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Camera Error", "Failed to grab frame.")
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            distances = face_recognition.face_distance(known_encodings, face_encoding)
            best_index = np.argmin(distances)
            best_distance = distances[best_index]

            if best_distance < 0.45:
                recognized_name = known_names[best_index]
                top, right, bottom, left = face_location
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, recognized_name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                if recognized_name.lower() == name.lower() and not already_logged_in:
                    log_login(recognized_name)
                    messagebox.showinfo("Login Successful", f"Login successful. Welcome, {recognized_name}!")
                    already_logged_in = True
            else:
                top, right, bottom, left = face_location
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                cv2.putText(frame, "Unknown", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        cv2.imshow("Login", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


# GUI setup
root = tk.Tk()
root.title("Binky Face Recognition")

# Set window size
window_width = 400
window_height = 340

# Get screen dimension
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Find the center point
center_x = int((screen_width - window_width) / 2)
center_y = int((screen_height - window_height) / 2)

# Set the position of the window to the center of the screen
root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
root.configure(bg="#ffe6f0")  # Light pink

# Title Label
title = tk.Label(root, text="Welcome 😊", font=("Comic Sans MS", 16, "bold"), bg="#ffe6f0", fg="#cc6699")
title.pack(pady=30)

# Name Entry
name_label = tk.Label(root, text="Enter your name:", font=("Comic Sans MS", 12), bg="#ffe6f0", fg="#cc6699")
name_label.pack(pady=5)

name_entry = tk.Entry(root, font=("Comic Sans MS", 12), width=25)
name_entry.pack(pady=5)

# Register Button
register_btn = tk.Button(root, text="Register", font=("Comic Sans MS", 12), bg="#ffb3d9", fg="white",
                         width=15, height=2, command=lambda: register_user(name_entry.get()))
register_btn.pack(pady=0)

# Login Button
login_btn = tk.Button(root, text="Login", font=("Comic Sans MS", 12), bg="#b3cde0", fg="white",
                      width=15, height=2, command=lambda: login_user(name_entry.get()))
login_btn.pack(pady=5)

# Footer
footer = tk.Label(root, text="Binky Security © 2025", font=("Comic Sans MS", 8), bg="#ffe6f0", fg="#999")
footer.pack(side="bottom", pady=10)

root.mainloop()

